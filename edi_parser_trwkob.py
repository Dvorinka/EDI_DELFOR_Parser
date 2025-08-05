import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDITrwkobParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI TRWKOB Parser")
        self.root.geometry("1200x800")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.setup_ui()
        self.main_window = None

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Styly pro tlačítka
        style = ttk.Style()
        style.configure('Excel.TButton', 
                      background='#217346',  # Excel zelená barva
                      foreground='white', 
                      font=('Segoe UI', 10, 'bold'),
                      padding=5)
        
        # Vytvoření tlačítek s odsazením a styly
        btn_back = ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main)
        btn_export = ttk.Button(btn_frame, 
                              text="📊 Export do Excelu", 
                              command=self.export_to_excel, 
                              style='Excel.TButton')
        
        # Uspořádání tlačítek s odsazením
        btn_back.pack(side=tk.LEFT, padx=(0, 5))
        btn_export.pack(side=tk.LEFT)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()

    def setup_info_tab(self):
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_delivery_tab(self):
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        columns = ('Datum od', 'Množství', 'Typ', 'SCC')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        # Set column widths
        self.delivery_tree.column('Datum od', width=100)
        self.delivery_tree.column('Množství', width=100)
        self.delivery_tree.column('Typ', width=150)
        self.delivery_tree.column('SCC', width=200)
        # Set column headings
        for col in columns:
            self.delivery_tree.heading(col, text=col)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    def setup_stats_tab(self):
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def parse_date(self, date_str, format_code):
        try:
            if format_code == '102':
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str
        except:
            return date_str

    def parse_edi_datetime(self, datetime_str):
        """Parsuje EDI datum/čas z UNB segmentu (YYMMDD:HHMM)"""
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str

    def load_file(self, filepath):
        """Load and parse the specified EDI file"""
        try:
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            self.parse_edi_file(content)
            self.display_data()
            return True
        except Exception as e:
            messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
            return False

    def parse_edi_file(self, content):
        lines = [line.strip() for line in content.strip().split("'") if line.strip()]
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        
        i = 0
        while i < len(lines):
            line = lines[i]
            if not line:
                i += 1
                continue
                
            # UNB - Interchange header
            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])
                i += 1
                    
            # BGM - Beginning of message
            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]
                i += 1
                            
            # NAD - Name and address
            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    code = parts[2]
                    name = parts[4] if len(parts) > 4 else ''
                    
                    # Process address parts
                    address_parts = []
                    for j in range(5, len(parts)):
                        if parts[j]:
                            address_parts.append(parts[j])
                    
                    full_address = ', '.join(address_parts) if address_parts else ''
                    
                    if role == 'BY':
                        self.partner_info['Kupující'] = name if name else code
                        if full_address:
                            self.partner_info['Kupující'] += f", {full_address}"
                    elif role == 'SE':
                        self.header_info['Příjemce'] = name if name else code
                        if full_address:
                            self.partner_info['Prodávající'] = f"{name if name else code}, {full_address}"
                        else:
                            self.partner_info['Prodávající'] = name if name else code
                    elif role == 'CN':
                        if full_address:
                            self.partner_info['Dodací adresa'] = f"{name if name else code}, {full_address}"
                        else:
                            self.partner_info['Dodací adresa'] = name if name else code
                i += 1
                        
            # LIN - Line item
            elif line.startswith('LIN'):
                parts = line.split('+')
                if len(parts) >= 4:
                    self.header_info['Číslo položky'] = parts[3]
                i += 1
                    
            # PIA - Product identification
            elif line.startswith('PIA'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Kód produktu'] = parts[2]
                i += 1
                
            # Handle delivery block (4 lines in specific order: QTY+113, SCC, DTM+63, DTM+64)
            elif line.startswith('QTY+113'):
                try:
                    # Create new delivery record
                    current_delivery = {}
                    
                    # 1. Process QTY+113 line (current line)
                    qty_parts = line.split('+')
                    if len(qty_parts) >= 2:
                        qty_info = qty_parts[1].split(':')
                        if len(qty_info) >= 3:
                            current_delivery['Množství'] = qty_info[1]
                            current_delivery['Jednotka'] = qty_info[2] if len(qty_info) > 2 else 'PCE'
                            current_delivery['Typ'] = 'Plánované množství'
                    
                    # Move to next line
                    i += 1
                    if i >= len(lines):
                        break
                        
                    # 2. Process SCC line (next line)
                    if i < len(lines) and lines[i].startswith('SCC'):
                        scc_parts = lines[i].split('+')
                        if len(scc_parts) >= 2:
                            current_delivery['SCC'] = scc_parts[1]
                        i += 1
                        
                        # 3. Process DTM+63 line (end date)
                        if i < len(lines) and lines[i].startswith('DTM+63'):
                            dtm_parts = lines[i].split('+')
                            if len(dtm_parts) >= 2:
                                dtm_info = dtm_parts[1].split(':')
                                if len(dtm_info) >= 3:
                                    current_delivery['Datum do'] = self.parse_date(dtm_info[1], dtm_info[2])
                            i += 1
                            
                            # 4. Process DTM+64 line (start date)
                            if i < len(lines) and lines[i].startswith('DTM+64'):
                                dtm_parts = lines[i].split('+')
                                if len(dtm_parts) >= 2:
                                    dtm_info = dtm_parts[1].split(':')
                                    if len(dtm_info) >= 3:
                                        current_delivery['Datum od'] = self.parse_date(dtm_info[1], dtm_info[2])
                                i += 1
                    
                    # Add completed delivery to schedules
                    if 'Množství' in current_delivery and 'Datum od' in current_delivery:
                        self.delivery_schedules.append(current_delivery.copy())
                
                except Exception as e:
                    print(f"Error processing delivery block: {e}")
                    i += 1  # Skip to next line on error
            else:
                # Skip any other lines we don't process
                i += 1

    def display_data(self):
        self.info_text.delete(1.0, tk.END)
        info_content = "=== HLAVIČKA DOKUMENTU ===\n"
        for key, value in self.header_info.items():
            if key != 'Příjemce_kód':
                info_content += f"{key}: {value}\n"
        if 'Příjemce' not in self.header_info and 'Příjemce_kód' in self.header_info:
            info_content += f"Příjemce: {self.header_info['Příjemce_kód']}\n"
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        self.info_text.insert(1.0, info_content)
        
        # Clear the tree
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)
            
        # Process all deliveries without deduplication
        deliveries_to_display = []
        
        # Filter out 'Maximální' and 'Minimální' types
        for delivery in self.delivery_schedules:
            delivery_type = delivery.get('Typ', '')
            # Skip 'Maximální' and 'Minimální' types
            if delivery_type in ['Maximální', 'Minimální']:
                continue
                
            date_from = delivery.get('Datum od', '')
            quantity = delivery.get('Množství', '')
            
            # Skip if any required field is missing
            if not (date_from and quantity):
                continue
                
            # Convert date strings to datetime objects for sorting
            try:
                date_obj = datetime.strptime(date_from, '%d.%m.%Y')
                deliveries_to_display.append((date_obj, delivery))
            except (ValueError, TypeError):
                # If date parsing fails, keep the original order
                deliveries_to_display.append((datetime.max, delivery))
        
        # Sort deliveries by date (from oldest to newest)
        deliveries_to_display.sort(key=lambda x: x[0])
        
        # Add all deliveries to the tree
        for date_obj, delivery in deliveries_to_display:
            scc_code = delivery.get('SCC', '')
            scc_desc = self.get_scc_description(scc_code)
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Datum od', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                scc_desc
            ))
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        total_qty = sum(int(d.get('Množství', 0)) for d in self.delivery_schedules if d.get('Množství', '').isdigit())
        stats_content += f"Celkové množství: {total_qty:,} kusů\n"
        type_stats = {}
        for delivery in self.delivery_schedules:
            delivery_type = delivery.get('Typ', 'Neznámý')
            if delivery_type not in type_stats:
                type_stats[delivery_type] = {'počet': 0, 'množství': 0}
            type_stats[delivery_type]['počet'] += 1
            if delivery.get('Množství', '').isdigit():
                type_stats[delivery_type]['množství'] += int(delivery.get('Množství', 0))
        stats_content += "\n=== STATISTIKY PODLE TYPU ===\n"
        for delivery_type, stats in type_stats.items():
            stats_content += f"{delivery_type}: {stats['počet']} dodávek, {stats['množství']:,} kusů\n"
        self.stats_text.insert(1.0, stats_content)

    def get_week_number(self, date_str):
        """Convert date string to ISO week number (WW)"""
        try:
            day, month, year = map(int, date_str.split('.'))
            dt = date(year, month, day)
            return dt.isocalendar()[1]
        except (ValueError, AttributeError):
            return ""
            
    def get_scc_description(self, scc_code):
        """Convert SCC code to descriptive name"""
        scc_mapping = {
            '10': 'Backlog',
            '1': 'Fix',
            '4': 'Forecast',
            '': 'Neznámé',
        }
        return scc_mapping.get(scc_code, f'Neznámý kód: {scc_code}')

    def export_to_excel(self):
        """Export delivery data to Excel with requested column order and sorting"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            # Process all deliveries
            processed_deliveries = []
            
            for delivery in self.delivery_schedules:
                date_str = delivery.get('Datum od', '')
                delivery_type = delivery.get('Typ', '')
                
                # Skip 'Maximální' and 'Minimální' types and empty dates
                if not date_str or delivery_type in ['Maximální', 'Minimální']:
                    continue
                    
                try:
                    # Parse date
                    date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
                    
                    # Get delivery details
                    quantity = delivery.get('Množství', '').strip("'")
                    scc_code = delivery.get('SCC', '')
                    item = delivery.get('Položka', '')  # Get item number if available
                    
                    # Store with item, date, and original delivery for sorting
                    processed_deliveries.append({
                        'item': item,
                        'date_obj': date_obj,
                        'date_str': date_str,
                        'quantity': quantity,
                        'type': delivery_type,
                        'scc_code': scc_code,
                        'scc_desc': self.get_scc_description(scc_code),
                        'delivery': delivery
                    })
                    
                except (ValueError, TypeError) as e:
                    print(f"Chyba při zpracování data: {date_str}, {e}")
                    continue
            
            # Sort by item and then by date
            processed_deliveries.sort(key=lambda x: (str(x['item'] or ''), x['date_obj']))
            
            # Create Excel workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Headers in requested order: datum, týden, množství, SCC, dodací místo
            headers = ["Datum", "Týden", "Množství", "SCC", "Dodací místo"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data
            row_num = 2
            for delivery_data in processed_deliveries:
                # 1. Datum (date) - formatted
                ws.cell(row=row_num, column=1, value=delivery_data['date_obj']).number_format = 'DD.MM.YYYY'
                
                # 2. Týden (week number) - as number
                ws.cell(row=row_num, column=2, value=delivery_data['date_obj'].isocalendar()[1])
                
                # 3. Množství (quantity) - as number
                try:
                    quantity = float(delivery_data['quantity']) if delivery_data['quantity'] else 0
                    ws.cell(row=row_num, column=3, value=quantity).number_format = '0'
                except (ValueError, TypeError):
                    ws.cell(row=row_num, column=3, value=delivery_data['quantity']).number_format = '0'
                
                # 4. SCC - as text
                ws.cell(row=row_num, column=4, value=delivery_data['scc_desc']).number_format = '@'
                
                # 5. Dodací místo (delivery address) - as text
                delivery_address = self.partner_info.get('Dodací adresa', '') or 'XTREME PRESSURE INJECTION JUAREZ, REC LOC 372, EL PASO, 79927'
                ws.cell(row=row_num, column=5, value=delivery_address).number_format = '@'
                
                row_num += 1

            # Apply number formatting to numeric columns
            for col in [2]:  # Only format week number column (quantity is already formatted)
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        # For dates, use the formatted string length
                        if hasattr(cell, 'is_date') and cell.is_date:
                            cell_value = cell.value.strftime('%d.%m.%Y') if cell.value else ''
                        else:
                            cell_value = str(cell.value) if cell.value is not None else ''
                        
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 30)

            # Save the file with trwkob in the name
            filename = f"dodavky_trwkob_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při exportu do Excelu: {str(e)}")

    def back_to_main(self):
        """Closes the current window and returns to the main application"""
        # Close current window
        self.root.destroy()
        # Return to main window
        if self.main_window:
            self.main_window.root.deiconify()  # Show the main window if it exists

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    app = EDITrwkobParser()
    app.run()
