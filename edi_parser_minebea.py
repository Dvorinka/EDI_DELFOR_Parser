import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDIDelforParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI MINEBEA Parser")
        self.root.geometry("1200x800")
        
        # Hlavní data
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        
        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.setup_ui()
        
        # If filepath was provided, load it automatically
        if filepath:
            self.load_file(filepath)
        
    def setup_ui(self):
        # Hlavní frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tlačítka pro ovládání
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Styly pro tlačítka
        style = ttk.Style()
        style.configure('Excel.TButton', 
                      background='#217346',  # Excel zelená barva
                      foreground='white', 
                      font=('Segoe UI', 10, 'bold'),
                      padding=5)
        
        # Vytvoření tlačítek s odsazením a styly
        btn_back = ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main)
        btn_export = ttk.Button(btn_frame, 
                              text="📊 Export do Excelu", 
                              command=self.export_to_excel, 
                              style='Excel.TButton')
        
        # Uspořádání tlačítek s odsazením
        btn_back.pack(side=tk.LEFT, padx=(0, 5))
        btn_export.pack(side=tk.LEFT)
        
        # Notebook pro záložky
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Záložka - Základní informace
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        
        # Záložka - Dodávky
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        
        # Záložka - Statistiky
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()
        
    def setup_info_tab(self):
        # Scrollable text widget pro základní informace
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def get_scc_description(self, scc_code):
        """Convert SCC code to descriptive name"""
        scc_mapping = {
            '10': 'Backlog',
            '1': 'Fix',
            '4': 'Forecast',
            '': 'Neznámé',
        }
        return scc_mapping.get(scc_code, f'Neznámý kód: {scc_code}')
        
    def setup_delivery_tab(self):
        # Treeview pro plán dodávek
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ('Datum od', 'Množství', 'Typ', 'SCC')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        
        # Definice sloupců
        for col in columns:
            self.delivery_tree.heading(col, text=col)
            self.delivery_tree.column(col, width=120)
        
        # Scrollbary
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
    def setup_stats_tab(self):
        # Statistiky
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
    def parse_date(self, date_str, format_code):
        """Parsuje datum podle EDI formátu"""
        try:
            if format_code == '203':  # CCYYMMDDHHMMSS
                # Return only date part without time
                return datetime.strptime(date_str, '%Y%m%d%H%M%S').strftime('%d.%m.%Y')
            elif format_code == '102':  # CCYYMMDD
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str.split(' ')[0]  # Return only date part if time is present
        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při parsování data {date_str} s formátem {format_code}: {e}")
            return date_str.split(' ')[0] if date_str else ''
    
    def parse_edi_datetime(self, datetime_str):
        """Parsuje EDI datum/čas z UNB segmentu (YYMMDD:HHMM)"""
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                # Přidáme 20 na začátek roku (předpokládáme 21. století)
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str
    
    def parse_edi_file(self, content):
        """Parsuje EDI DELFOR soubor"""
        lines = [line.strip() for line in content.strip().split("'") if line.strip()]
        
        # Reset dat
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        
        i = 0
        while i < len(lines):
            line = lines[i]
            if not line:
                i += 1
                continue
                
            # UNB - Interchange header
            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])
                i += 1
                    
            # BGM - Beginning of message
            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]
                i += 1
                            
            # NAD - Name and address
            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    code = parts[2] if len(parts) > 2 else ''
                    name = parts[4] if len(parts) > 4 else ''
                    
                    # Process address parts
                    address_parts = []
                    for j in range(5, len(parts)):
                        if parts[j]:
                            address_parts.append(parts[j])
                    
                    full_address = ', '.join(address_parts) if address_parts else ''
                    
                    if role == 'BY':
                        self.partner_info['Kupující'] = name
                        if full_address:
                            self.partner_info['Kupující'] += f", {full_address}"
                    elif role == 'SE':
                        if code == self.header_info.get('Příjemce_kód', ''):
                            self.header_info['Příjemce'] = name
                        
                        if full_address:
                            self.partner_info['Prodávající'] = f"{name}, {full_address}"
                        else:
                            self.partner_info['Prodávající'] = name
                    elif role == 'CN':
                        if full_address:
                            self.partner_info['Dodací adresa'] = f"{name}, {full_address}"
                        else:
                            self.partner_info['Dodací adresa'] = name
                i += 1
                        
            # LIN - Line item
            elif line.startswith('LIN'):
                parts = line.split('+')
                if len(parts) >= 4:
                    self.header_info['Číslo položky'] = parts[3]
                i += 1
                    
            # PIA - Product identification
            elif line.startswith('PIA'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Kód produktu'] = parts[2]
                i += 1
                
            # Handle delivery block (4 lines in specific order: QTY+113, SCC, DTM+63, DTM+64)
            elif line.startswith('QTY+113'):
                try:
                    # Create new delivery record
                    current_delivery = {}
                    
                    # 1. Process QTY+113 line (current line)
                    qty_parts = line.split('+')
                    if len(qty_parts) >= 2:
                        qty_info = qty_parts[1].split(':')
                        if len(qty_info) >= 3:
                            current_delivery['Množství'] = qty_info[1]
                            current_delivery['Jednotka'] = qty_info[2]
                            current_delivery['Typ'] = 'Plánované množství'
                    
                    # Move to next line
                    i += 1
                    if i >= len(lines):
                        break
                        
                    # 2. Process SCC line (next line)
                    if lines[i].startswith('SCC'):
                        scc_parts = lines[i].split('+')
                        if len(scc_parts) >= 2:
                            current_delivery['SCC'] = scc_parts[1]
                        i += 1
                        
                        # 3. Process DTM+63 line (end date)
                        if i < len(lines) and lines[i].startswith('DTM+63'):
                            dtm_parts = lines[i].split('+')
                            if len(dtm_parts) >= 2:
                                dtm_info = dtm_parts[1].split(':')
                                if len(dtm_info) >= 3:
                                    current_delivery['Datum do'] = self.parse_date(dtm_info[1], dtm_info[2])
                            i += 1
                            
                            # 4. Process DTM+64 line (start date)
                            if i < len(lines) and lines[i].startswith('DTM+64'):
                                dtm_parts = lines[i].split('+')
                                if len(dtm_parts) >= 2:
                                    dtm_info = dtm_parts[1].split(':')
                                    if len(dtm_info) >= 3:
                                        current_delivery['Datum od'] = self.parse_date(dtm_info[1], dtm_info[2])
                                i += 1
                    
                    # Add completed delivery to schedules
                    if 'Množství' in current_delivery and 'Datum od' in current_delivery:
                        self.delivery_schedules.append(current_delivery.copy())
                
                except Exception as e:
                    print(f"Error processing delivery block: {e}")
                    i += 1  # Skip to next line on error
            else:
                # Skip any other lines we don't process
                i += 1
    
    def load_file(self, filepath):
        """Načte EDI soubor"""
        try:
            # Check if the window still exists
            if not hasattr(self, 'root') or not self.root.winfo_exists():
                return False
                
            with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
                
            self.parse_edi_file(content)
            
            # Check again before updating UI
            if hasattr(self, 'root') and self.root.winfo_exists():
                self.display_data()
                return True
            return False
            
        except Exception as e:
            # Safely show error if window still exists
            if hasattr(self, 'root') and self.root.winfo_exists():
                messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
            return False
    
    def display_data(self):
        """Zobrazí naparsovaná data"""
        # Check if window still exists
        if not hasattr(self, 'info_text') or not hasattr(self, 'root') or not self.root.winfo_exists():
            return
            
        try:
            # Základní informace
            self.info_text.delete(1.0, tk.END)
            info_content = "=== HLAVIČKA DOKUMENTU ===\n"
            for key, value in self.header_info.items():
                # Přeskočíme pomocný klíč
                if key != 'Příjemce_kód':
                    info_content += f"{key}: {value}\n"
            
            # Pokud nemáme název příjemce, zobrazíme alespoň kód
            if 'Příjemce' not in self.header_info and 'Příjemce_kód' in self.header_info:
                info_content += f"Příjemce: {self.header_info['Příjemce_kód']}\n"
        except Exception as e:
            # Skip if there's an error during display
            messagebox.showerror("Chyba", f"Chyba při zobrazování dat: {e}")
            return
        
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        
        self.info_text.insert(1.0, info_content)
        
        # Plán dodávek
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)
            
        for delivery in self.delivery_schedules:
            scc_code = delivery.get('SCC', '')
            scc_desc = self.get_scc_description(scc_code)
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Datum od', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                scc_desc
            ))
        
        # Statistiky
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        
        total_qty = sum(int(d.get('Množství', 0)) for d in self.delivery_schedules if d.get('Množství', '').isdigit())
        stats_content += f"Celkové množství: {total_qty:,} kusů\n"
        
        # Statistiky podle typu
        type_stats = {}
        for delivery in self.delivery_schedules:
            delivery_type = delivery.get('Typ', 'Neznámý')
            if delivery_type not in type_stats:
                type_stats[delivery_type] = {'počet': 0, 'množství': 0}
            type_stats[delivery_type]['počet'] += 1
            if delivery.get('Množství', '').isdigit():
                type_stats[delivery_type]['množství'] += int(delivery.get('Množství', 0))
        
        stats_content += "\n=== STATISTIKY PODLE TYPU ===\n"
        for delivery_type, stats in type_stats.items():
            stats_content += f"{delivery_type}: {stats['počet']} dodávek, {stats['množství']:,} kusů\n"
        
        self.stats_text.insert(1.0, stats_content)
    
    def get_week_number(self, date_str):
        """Převede řetězec s datem na číslo kalendářního týdne (WW)"""
        if not date_str:
            return ""
        try:
            # Handle case where time might be included
            date_part = date_str.split(' ')[0]
            day, month, year = map(int, date_part.split('.'))
            # Handle 2-digit year
            if year < 100:
                year += 2000  # Assuming 21st century for 2-digit years
            dt = date(year, month, day)
            week_num = dt.isocalendar()[1]
            return week_num
        except Exception as e:
            # Log error silently
            return ""

    def export_to_excel(self):
        """Exportuje data o dodávkách do Excelu s kalendářními týdny"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Hlavičky podle požadavku: datum, týden, množství, SCC, zbytek ad lib
            headers = ["Datum", "Týden", "Množství", "SCC", "Dodací místo"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Připravíme data pro řazení
            prepared_data = []
            for delivery in self.delivery_schedules:
                # Získáme položku (item) - pokud neexistuje, použijeme prázdný řetězec
                item = delivery.get('Položka', '')
                
                # Zpracujeme datum pro řazení
                date_str = delivery.get('Datum od', '')
                date_for_sort = None
                if date_str:
                    try:
                        date_parts = date_str.split(' ')[0].split('.')
                        if len(date_parts) == 3:
                            date_for_sort = datetime(int(date_parts[2]), int(date_parts[1]), int(date_parts[0]))
                    except (ValueError, IndexError):
                        pass
                
                prepared_data.append({
                    'item': item,
                    'date_for_sort': date_for_sort or datetime.max,
                    'delivery': delivery
                })
            
            # Seřadíme data podle položky a data
            prepared_data.sort(key=lambda x: (x['item'] or '', x['date_for_sort']))

            # Data
            row_num = 2
            for item_data in prepared_data:
                delivery = item_data['delivery']
                date_from = delivery.get('Datum od', '')
                week_num = self.get_week_number(date_from) if date_from else ""
                scc_code = delivery.get('SCC', '')
                scc_desc = self.get_scc_description(scc_code)
                
                # Datum (sloupec 1)
                try:
                    if date_from:
                        # Remove time part if present
                        date_from = date_from.split(' ')[0]
                        date_from_obj = datetime.strptime(date_from, '%d.%m.%Y')
                        ws.cell(row=row_num, column=1, value=date_from_obj).number_format = 'DD.MM.YYYY'
                except Exception as e:
                    ws.cell(row=row_num, column=2, value=date_from.split(' ')[0] if date_from else '')
                
                # Týden (sloupec 2)
                try:
                    week_num = int(week_num) if week_num else 0
                    ws.cell(row=row_num, column=2, value=week_num).number_format = '0'
                except (ValueError, TypeError):
                    ws.cell(row=row_num, column=3, value=0)
                
                # Množství (sloupec 3)
                quantity = delivery.get('Množství', '')
                try:
                    if isinstance(quantity, str):
                        quantity = quantity.strip("'")
                        qty_value = float(quantity) if quantity else 0.0
                    else:
                        qty_value = float(quantity) if quantity is not None else 0.0
                    ws.cell(row=row_num, column=3, value=qty_value).number_format = '0'
                except (ValueError, TypeError):
                    ws.cell(row=row_num, column=4, value=0.0).number_format = '0'
                
                # SCC (sloupec 4)
                scc_desc = self.get_scc_description(str(delivery.get('SCC', '')))
                ws.cell(row=row_num, column=4, value=str(scc_desc)).number_format = '@'
                
                # Dodací místo (sloupec 5)
                ws.cell(row=row_num, column=5, value=str(self.partner_info.get('Dodací adresa', '') or 'XTREME PRESSURE INJECTION JUAREZ, REC LOC 372, EL PASO, 79927')).number_format = '@'
                
                row_num += 1

            # Apply number formatting to numeric columns
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Format quantity column (column 2) as number with no decimal places
                if row[2].value is not None:  # Column 3 (0-based index 2)
                    row[2].number_format = '0'
                # Format week number column (column 1) as number with no decimal places
                if row[1].value is not None:  # Column 2 (0-based index 1)
                    row[1].number_format = '0'
            
            # Automatické přizpůsobení šířky sloupců
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        # For dates, use the formatted string length
                        if hasattr(cell, 'is_date') and cell.is_date:
                            cell_value = cell.value.strftime('%d.%m.%Y') if cell.value else ''
                        else:
                            cell_value = str(cell.value) if cell.value is not None else ''
                        
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 30)

            # Uložení souboru
            filename = f"dodavky_minebea_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Chyba", f"Chyba při exportu do Excelu: {str(e)}")

    def on_closing(self):
        """Handle window close event"""
        self.root.destroy()  # Close the current window
        
    def back_to_main(self):
        """Closes the current window"""
        self.root.destroy()
    
    def run(self):
        """Spustí aplikaci"""
        self.root.mainloop()

if __name__ == "__main__":
    app = EDIDelforParser()
    app.run()