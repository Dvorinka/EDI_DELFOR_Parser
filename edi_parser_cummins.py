import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class EDIDelforCumminsParser:
    def __init__(self, filepath=None):
        self.root = tk.Tk()
        self.root.title("EDI Cummins Parser")
        self.root.geometry("1200x800")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.line_items = []
        
        # Handle window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.setup_ui()
        if filepath:
            self.load_file(filepath)

    def setup_ui(self):
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        # Style configuration for buttons
        style = ttk.Style()
        style.configure('Excel.TButton', 
                       background='#217346',  # Excel green color
                       foreground='white', 
                       font=('Segoe UI', 10, 'bold'),
                       padding=5)
        
        # Add buttons with padding and styling
        btn_back = ttk.Button(btn_frame, text="Zpět na hlavní okno", command=self.back_to_main)
        btn_export = ttk.Button(btn_frame, 
                              text="📊 Export do Excelu", 
                              command=self.export_to_excel, 
                              style='Excel.TButton')
        
        # Pack buttons with padding
        btn_back.pack(side=tk.LEFT, padx=(0, 5))
        btn_export.pack(side=tk.LEFT)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        self.info_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.info_frame, text="Základní informace")
        self.delivery_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.delivery_frame, text="Plán dodávek")
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="Statistiky")
        self.setup_info_tab()
        self.setup_delivery_tab()
        self.setup_stats_tab()

    def setup_info_tab(self):
        text_frame = ttk.Frame(self.info_frame)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.info_text = tk.Text(text_frame, wrap=tk.WORD, font=('Courier', 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=scrollbar.set)
        self.info_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_delivery_tab(self):
        tree_frame = ttk.Frame(self.delivery_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        # Removed 'Jednotka' column as requested
        columns = ('Položka', 'Popis', 'Datum', 'Množství', 'Typ', 'SCC', 'Release')
        self.delivery_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        for col in columns:
            self.delivery_tree.heading(col, text=col)
            if col == 'Popis':
                self.delivery_tree.column(col, width=200)
            elif col == 'Položka':
                self.delivery_tree.column(col, width=100)
            else:
                self.delivery_tree.column(col, width=80)
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.delivery_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.delivery_tree.xview)
        self.delivery_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.delivery_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    def setup_stats_tab(self):
        stats_frame = ttk.Frame(self.stats_frame)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.stats_text = tk.Text(stats_frame, wrap=tk.WORD, font=('Courier', 10))
        stats_scrollbar = ttk.Scrollbar(stats_frame, orient=tk.VERTICAL, command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        self.stats_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        stats_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def parse_date(self, date_str, format_code):
        try:
            if format_code == '102':
                return datetime.strptime(date_str, '%Y%m%d').strftime('%d.%m.%Y')
            else:
                return date_str
        except:
            return date_str

    def parse_edi_datetime(self, datetime_str):
        try:
            if ':' in datetime_str:
                date_part, time_part = datetime_str.split(':')
                full_date = '20' + date_part
                formatted_date = datetime.strptime(full_date, '%Y%m%d').strftime('%d.%m.%Y')
                formatted_time = datetime.strptime(time_part, '%H%M').strftime('%H:%M')
                return f"{formatted_date} {formatted_time}"
            return datetime_str
        except:
            return datetime_str

    def get_scc_description(self, scc_code):
        scc_map = {
            '10': 'Backlog',
            '1': 'Firm',
            '4': 'Forecast'
        }
        return scc_map.get(scc_code, f'{scc_code}')

    def parse_edi_file(self, content):
        lines = content.strip().split("'")
        self.header_info = {}
        self.partner_info = {}
        self.delivery_schedules = []
        self.line_items = []
        
        # Current parsing state
        current_part_number = ''
        current_description = ''
        current_location = ''
        current_po = ''
        current_scc = ''
        current_release = ''
        
        # Track current line item details
        current_line_item = None
        
        # Temporary storage for quantity waiting for date
        pending_quantities = []
        
        def create_or_update_line_item():
            nonlocal current_line_item
            if not current_part_number:
                return None
                
            line_item = next((item for item in self.line_items 
                           if item['Položka'] == current_part_number), None)
            
            if not line_item:
                line_item = {
                    'Položka': current_part_number,
                    'Popis': current_description,
                    'Objednávka': current_po,
                    'Lokace': current_location,
                    'RFF': {}
                }
                self.line_items.append(line_item)
                
            # Update current line item reference
            current_line_item = line_item
            return line_item

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith('UNB'):
                parts = line.split('+')
                if len(parts) >= 5:
                    self.header_info['Odesílatel'] = parts[2]
                    self.header_info['Příjemce_kód'] = parts[3]
                    self.header_info['Datum/Čas'] = self.parse_edi_datetime(parts[4])

            elif line.startswith('UNH'):
                parts = line.split('+')
                if len(parts) >= 2:
                    self.header_info['ID zprávy'] = parts[1]

            elif line.startswith('BGM'):
                parts = line.split('+')
                if len(parts) >= 3:
                    self.header_info['Číslo zprávy'] = parts[2]

            elif line.startswith('DTM'):
                parts = line.split('+')
                if len(parts) >= 2:
                    dtm_parts = parts[1].split(':')
                    if len(dtm_parts) >= 3:
                        code = dtm_parts[0]
                        value = dtm_parts[1]
                        fmt = dtm_parts[2]
                        formatted_date = self.parse_date(value, fmt)
                        if code == '137':
                            self.header_info['Datum dokumentu'] = formatted_date
                        elif code == '2':
                            # This is a delivery date - match with pending quantities
                            # Only create entries if we have quantities to process
                            if pending_quantities:
                                # For SCC 10 (Backlog), we only take the first quantity
                                if current_scc == '10' and len(pending_quantities) > 0:
                                    qty_info = pending_quantities[0]
                                    # Create line item if it doesn't exist
                                    line_item = next((item for item in self.line_items if item['Položka'] == current_part_number), None)
                                    if not line_item:
                                        line_item = {
                                            'Položka': current_part_number,
                                            'Popis': current_description,
                                            'Objednávka': current_po,
                                            'Lokace': current_location
                                        }
                                        self.line_items.append(line_item)
                                    
                                    delivery = {
                                        'Položka': current_part_number,
                                        'Popis': current_description,
                                        'Datum': formatted_date,
                                        'Množství': qty_info['quantity'],
                                        'Typ': qty_info['type'],
                                        'SCC': self.get_scc_description(current_scc),
                                        'Release': current_release,
                                        'Objednávka': current_po
                                    }
                                    self.delivery_schedules.append(delivery)
                                else:
                                    # For other SCCs, process all quantities
                                    for qty_info in pending_quantities:
                                        delivery = {
                                            'Položka': current_part_number,
                                            'Popis': current_description,
                                            'Datum': formatted_date,
                                            'Množství': qty_info['quantity'],
                                            'Typ': qty_info['type'],
                                            'SCC': self.get_scc_description(current_scc),
                                            'Release': current_release
                                        }
                                        self.delivery_schedules.append(delivery)
                            pending_quantities.clear()
                            # Don't reset release here to maintain it for next entries

            elif line.startswith('NAD'):
                parts = line.split('+')
                if len(parts) >= 3:
                    role = parts[1]
                    if role == 'SU':  # Supplier
                        name_parts = [p.replace('?+', '').replace('?', '').strip() for p in parts[4:] if p]
                        self.partner_info['Dodavatel'] = ' '.join(name_parts)
                    elif role == 'ST':  # Ship To
                        name_parts = [p.replace('?+', '').replace('?', '').strip() for p in parts[4:] if p]
                        self.partner_info['Příjemce'] = ' '.join(name_parts)
                        # Store the full address for delivery location
                        if len(parts) > 5:  # If there are address components
                            address_parts = []
                            # Get address lines (parts[5] and beyond)
                            for part in parts[5:]:
                                if ':' in part:  # Skip parts with qualifiers
                                    break
                                address_parts.append(part.replace('?+', '').replace('?', '').strip())
                            if address_parts:
                                self.partner_info['Dodací adresa'] = ', '.join(address_parts)
                                # If no specific address found, use the recipient name as fallback
                                if not self.partner_info['Dodací adresa'] and name_parts:
                                    self.partner_info['Dodací adresa'] = ' '.join(name_parts)

            elif line.startswith('LIN'):
                # Save previous line item if it exists
                if current_part_number:
                    # Process previous line item if exists
                    create_or_update_line_item()
                
                parts = line.split('+')
                # Process LIN segment
                
                if len(parts) >= 4:
                    # Reset part information for new line item
                    current_part_number = ''
                    current_description = ''
                    current_scc = ''
                    current_release = ''
                    current_line_item = None
                    pending_quantities = []
                    
                    # Try to find part number in the LIN segment
                    for i, part in enumerate(parts[3:], 3):  # Skip the first 3 parts (LIN, line number, action code)
                        # Process part
                        if ':' in part:  # If the part contains a colon, it might be a part number
                            part_info = part.split(':')
                            # Process part info
                            if len(part_info) >= 2 and part_info[1] == 'IN':  # Look for part number with 'IN' qualifier
                                current_part_number = part_info[0]
                                # Found part number with IN qualifier
                                break
                            elif not current_part_number:  # If no 'IN' qualifier found, take the first part
                                current_part_number = part_info[0]
                                # Using first part as part number
                    
                    # If still no part number found, try to get it from the last part
                    if not current_part_number and parts[3:]:
                        current_part_number = parts[3].split(':')[0]
                        # Using fallback part number
                    
                    # Final part number processed

            elif line.startswith('IMD'):
                parts = line.split('+')
                if len(parts) >= 4:
                    # Extract item description - fixed to properly handle the format
                    # Looking for the 4th element which contains the description
                    desc_part = parts[3] if len(parts) > 3 else ''
                    
                    # Remove leading colons and extract the actual description
                    if desc_part.startswith(':::'):
                        current_description = desc_part[3:].strip()
                    elif desc_part.startswith('::'):
                        current_description = desc_part[2:].strip()
                    elif desc_part.startswith(':'):
                        current_description = desc_part[1:].strip()
                    else:
                        current_description = desc_part.strip()
                    
                    # Clean up any remaining formatting
                    current_description = current_description.replace(':', '').strip()

            elif line.startswith('LOC'):
                parts = line.split('+')
                if len(parts) >= 3:
                    current_location = parts[2]

            elif line.startswith('RFF'):
                parts = line.split('+')
                # Process RFF segment
                
                if len(parts) >= 2:
                    ref_parts = parts[1].split(':')
                    if len(ref_parts) >= 2:
                        ref_type = ref_parts[0]
                        ref_value = ref_parts[1]
                        
                        # Found RFF reference
                        
                        # Create or update line item if it doesn't exist
                        if not current_line_item:
                            # Create new line item if none exists
                            create_or_update_line_item()
                        
                        # Store the reference in the current line item
                        if current_line_item:
                            if 'RFF' not in current_line_item:
                                current_line_item['RFF'] = {}
                            current_line_item['RFF'][ref_type] = ref_value
                            # RFF stored in line item
                            
                            # Special handling for order numbers
                            if ref_type == 'ON':
                                current_po = ref_value
                                current_line_item['Objednávka'] = current_po
                                # Order number set
                            elif ref_type == 'RE':
                                current_release = ref_value
                                # Clear any pending quantities to ensure release number is applied to new quantities
                                pending_quantities = []
                                # Release number set
                        else:
                            # No line item available for RFF
                            pass

            elif line.startswith('SCC'):
                parts = line.split('+')
                if len(parts) >= 2:
                    current_scc = parts[1]
                    # Clear pending quantities when new SCC starts to prevent duplicates
                    pending_quantities = []
                    # Only reset release for backlog (SCC 10)
                    if current_scc == '10':
                        current_release = ''

            elif line.startswith('QTY'):
                parts = line.split('+')
                if len(parts) >= 2:
                    qty_parts = parts[1].split(':')
                    if len(qty_parts) >= 2:
                        qty_type = qty_parts[0]
                        quantity = qty_parts[1]
                        # Removed unit extraction as we don't need it
                        
                        # Determine quantity type
                        qty_type_desc = 'Neznámý'
                        if qty_type == '1':
                            qty_type_desc = 'Dodávka'
                        elif qty_type == '3':
                            qty_type_desc = 'Kumulativní'
                        elif qty_type == '48':
                            qty_type_desc = 'Plánované'
                        
                        # Store quantity info waiting for corresponding date
                        pending_quantities.append({
                            'quantity': quantity,
                            'type': qty_type_desc
                        })

        # Store line items for reference
        unique_parts = {}
        for delivery in self.delivery_schedules:
            part_num = delivery['Položka']
            if part_num not in unique_parts:
                unique_parts[part_num] = {
                    'Položka': part_num,
                    'Popis': delivery['Popis']
                }
        
        self.line_items = list(unique_parts.values())

    def load_file(self, filepath=None):
        if filepath:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.parse_edi_file(content)
                self.display_data()
                return True
            except Exception as e:
                messagebox.showerror("Chyba", f"Nelze načíst soubor: {str(e)}")
                return False
        return False

    def display_data(self):
        # Display header info
        self.info_text.delete(1.0, tk.END)
        info_content = "=== HLAVIČKA DOKUMENTU ===\n"
        for key, value in self.header_info.items():
            if key != 'Příjemce_kód':
                info_content += f"{key}: {value}\n"
        info_content += "\n=== INFORMACE O PARTNERECH ===\n"
        for key, value in self.partner_info.items():
            info_content += f"{key}: {value}\n"
        self.info_text.insert(1.0, info_content)

        # Display delivery schedules
        for item in self.delivery_tree.get_children():
            self.delivery_tree.delete(item)

        # Sort deliveries by date
        def date_sort_key(delivery):
            date_str = delivery.get('Datum', '')
            try:
                return datetime.strptime(date_str, '%d.%m.%Y') if date_str else datetime.max
            except:
                return datetime.max

        sorted_deliveries = sorted(self.delivery_schedules, key=date_sort_key)
        
        for delivery in sorted_deliveries:
            self.delivery_tree.insert('', tk.END, values=(
                delivery.get('Položka', ''),
                delivery.get('Popis', ''),
                delivery.get('Datum', ''),
                delivery.get('Množství', ''),
                delivery.get('Typ', ''),
                delivery.get('SCC', ''),
                delivery.get('Release', '')
            ))

        # Display statistics
        self.stats_text.delete(1.0, tk.END)
        stats_content = "=== STATISTIKY ===\n"
        stats_content += f"Celkový počet dodávek: {len(self.delivery_schedules)}\n"
        stats_content += f"Počet různých položek: {len(self.line_items)}\n"
        
        # Group by SCC
        scc_stats = {}
        total_qty = 0
        for delivery in self.delivery_schedules:
            scc = delivery.get('SCC', 'Neznámý')
            qty_str = delivery.get('Množství', '0')
            try:
                qty = int(qty_str)
                total_qty += qty
                if scc not in scc_stats:
                    scc_stats[scc] = {'count': 0, 'total_qty': 0}
                scc_stats[scc]['count'] += 1
                scc_stats[scc]['total_qty'] += qty
            except:
                pass
        
        stats_content += f"Celkové množství: {total_qty:,} kusů\n\n"
        stats_content += "=== STATISTIKY PO SCC ===\n"
        for scc, stats in scc_stats.items():
            stats_content += f"{scc}: {stats['count']} dodávek, {stats['total_qty']:,} kusů\n"
        
        self.stats_text.insert(1.0, stats_content)

    def on_closing(self):
        """Handle window close event"""
        self.root.destroy()  # Close the current window
        
    def back_to_main(self):
        """Closes the current window"""
        self.root.destroy()

    def get_week_number(self, date_str):
        """Convert date string to ISO week number"""
        try:
            # Handle different date formats
            if '.' in date_str:
                date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
            else:
                date_obj = datetime.strptime(date_str, '%Y%m%d').date()
            return date_obj.isocalendar()[1]  # Returns ISO week number
        except Exception as e:
            # Log error silently
            return ""

    def export_to_excel(self):
        """Export delivery data to Excel with calendar weeks, color-coded by part"""
        if not self.delivery_schedules:
            messagebox.showwarning("Upozornění", "Žádná data k exportu")
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dodávky"

            # Get unique part numbers and assign colors
            unique_parts = list(set([item.get('Položka', '') for item in self.delivery_schedules if item.get('Položka')]))
            # Generate distinct colors for each part
            colors = [
                'FFE6B8', 'B8D1E6', 'E6B8B8', 'B8E6C3', 'E6D5B8',
                'D1B8E6', 'B8E6E6', 'E6B8D1', 'B8C3E6', 'E6E6B8',
                'B8E6D1', 'E6B8E6', 'B8E6B8', 'E6C3B8', 'B8D1E6',
                'E6B8C3', 'B8E6D9', 'E6B8D9', 'B8E6B8', 'E6B8FF'
            ]
            part_colors = {}
            for i, part in enumerate(unique_parts):
                part_colors[part] = colors[i % len(colors)]

            # Headers in requested order: položka, datum, týden, množství, SCC, zbytek ad lib
            headers = ["Položka", "Datum", "Týden", "Množství", "SCC", "Dodací místo"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
            # Add legend headers
            legend_headers = ["Legenda:", "Položka", "Popis"]
            for col_num, header in enumerate(legend_headers, 10):  # Start from column J
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Prepare data for sorting by item and date
            prepared_data = []
            for item in self.delivery_schedules:
                # Get item number (Položka)
                part_number = item.get('Položka', '')
                
                # Parse date for sorting
                date_str = item.get('Datum', '')
                date_for_sort = None
                if date_str:
                    try:
                        if '.' in date_str:
                            date_parts = date_str.split('.')
                            if len(date_parts) == 3:
                                date_for_sort = datetime(int(date_parts[2]), int(date_parts[1]), int(date_parts[0]))
                        else:
                            # Handle YYYYMMDD format if needed
                            date_for_sort = datetime.strptime(date_str, '%Y%m%d')
                    except (ValueError, IndexError):
                        pass
                
                prepared_data.append({
                    'part_number': part_number,
                    'date_for_sort': date_for_sort or datetime.max,
                    'item': item
                })
            
            # Sort by item and date
            prepared_data.sort(key=lambda x: (str(x['part_number'] or ''), x['date_for_sort']))

            # Add data to worksheet
            row_num = 2
            for data in prepared_data:
                item = data['item']
                
                # Get week number from date
                week_num = self.get_week_number(item.get('Datum', ''))
                
                # Format quantity as number, removing any leading quotes
                quantity = item.get('Množství', '')
                if isinstance(quantity, str):
                    quantity = quantity.strip("'")
                    try:
                        quantity = float(quantity) if quantity else 0
                    except (ValueError, TypeError):
                        quantity = 0
                
                # Format week number as number
                try:
                    week_num = int(week_num) if week_num else 0
                except (ValueError, TypeError):
                    week_num = 0
                
                # Get part number
                part_number = item.get('Položka', '')
                
                # Get SCC description
                scc = item.get('SCC', '')
                scc_desc = self.get_scc_description(str(scc)) if scc else ''
                
                # Get delivery location
                delivery_location = str(self.partner_info.get('Dodací adresa', '') or '')
                if not delivery_location.strip():
                    delivery_location = 'Cummins Inc., 500 Jackson Street, Columbus, IN 47201, USA'  # Default Cummins address
                
                # Get part description for legend
                part_description = item.get('Popis', '')
                
                # Get color for this part
                part_color = part_colors.get(part_number, 'FFFFFF')  # Default to white if part not found
                
                # 1. Položka (as text with colored background and part number as text)
                cell = ws.cell(row=row_num, column=1, value=str(part_number))
                cell.number_format = '@'
                cell.fill = openpyxl.styles.PatternFill(start_color=part_color, end_color=part_color, fill_type='solid')
                cell.font = Font(color='000000')  # Ensure text is black for visibility
                
                # 2. Datum (formatted date) - not colored
                date_str = item.get('Datum', '')
                try:
                    if date_str:
                        if '.' in date_str:
                            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                        else:
                            date_obj = datetime.strptime(date_str, '%Y%m%d')
                        cell = ws.cell(row=row_num, column=2, value=date_obj)
                        cell.number_format = 'DD.MM.YYYY'
                    else:
                        cell = ws.cell(row=row_num, column=2, value='')
                except:
                    cell = ws.cell(row=row_num, column=2, value=date_str)
                
                # 3. Týden (week number) - not colored
                cell = ws.cell(row=row_num, column=3, value=week_num)
                cell.number_format = '0'
                
                # 4. Množství (quantity as number) - not colored
                try:
                    if isinstance(quantity, (int, float)):
                        qty_value = float(quantity)
                    else:
                        qty_str = str(quantity).strip().replace("'", "")
                        qty_value = float(qty_str) if qty_str.replace('.', '', 1).isdigit() else 0.0
                    cell = ws.cell(row=row_num, column=4, value=qty_value)
                    cell.number_format = '0'
                except (ValueError, AttributeError):
                    cell = ws.cell(row=row_num, column=4, value=0.0)
                    cell.number_format = '0'
                
                # 5. SCC (as text) - not colored
                cell = ws.cell(row=row_num, column=5, value=str(scc_desc))
                cell.number_format = '@'
                
                # 6. Dodací místo (delivery address as text) - not colored
                cell = ws.cell(row=row_num, column=6, value=delivery_location)
                cell.number_format = '@'
                
                row_num += 1

            # Apply number formatting to numeric columns
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Skip empty rows or rows with insufficient columns
                if len(row) < 6:  # We need at least 6 columns (0-5)
                    continue
                    
                # Format week number (column 1) as number with no decimal places
                if row[0].value is not None and row[0].value != '':  # Column 1 (0-based index 0)
                    if isinstance(row[0].value, (int, float)):
                        row[0].number_format = '0'
                
                # Format part number (column 3) as number with no decimal places if it's a number
                if len(row) > 2 and row[2].value is not None and row[2].value != '':  # Column 3 (0-based index 2)
                    if isinstance(row[2].value, (int, float)):
                        row[2].number_format = '0'
                
                # Format quantity (column 4) as number with no decimal places
                if len(row) > 4 and row[4].value is not None and row[4].value != '':  # Column 5 (0-based index 4)
                    if isinstance(row[4].value, (int, float)):
                        row[4].number_format = '0'
            
            # Add legend headers (only in columns J and K)
            ws.cell(row=1, column=10, value="Legenda:").font = Font(bold=True)
            ws.cell(row=1, column=11, value="Popis").font = Font(bold=True)
            # Clear any existing header in column L
            if ws.cell(row=1, column=12).value == "Popis":
                ws.cell(row=1, column=12, value="")
            
            # Add legend items starting from row 2
            legend_row = 2
            for part_number, color in part_colors.items():
                # Get part description
                part_description = ''
                for item in self.delivery_schedules:
                    if item.get('Položka') == part_number:
                        part_description = item.get('Popis', '')
                        break
                
                # Set column J width to 0.75 inches (approximately 8.43 units in Excel)
                ws.column_dimensions['J'].width = 10
                
                # Create a cell with colored background and part number as text
                legend_cell = ws.cell(row=legend_row, column=10, value=str(part_number))
                legend_cell.fill = openpyxl.styles.PatternFill(
                    start_color=color, end_color=color, fill_type='solid')
                legend_cell.font = Font(color='000000', bold=True)  # Black text, bold
                legend_cell.alignment = Alignment(horizontal='center')
                
                # Part description in next column
                ws.cell(row=legend_row, column=11, value=part_description)
                
                legend_row += 1
            
            # Auto-adjust column widths for all columns
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                
                # Skip the color swatch column (J) for width adjustment
                if column_letter == 'J':
                    ws.column_dimensions[column_letter].width = 10  # Fixed width for color swatch (0.75")
                    continue
                    
                for cell in col:
                    try:
                        # For dates, use the formatted string length
                        if hasattr(cell, 'is_date') and cell.is_date:
                            cell_value = cell.value.strftime('%d.%m.%Y') if cell.value else ''
                        else:
                            cell_value = str(cell.value) if cell.value is not None else ''
                        
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                
                # Set a reasonable maximum width to prevent extremely wide columns
                adjusted_width = min((max_length + 2), 30)
                
                # Set minimum width for better readability
                if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:  # Main data columns
                    adjusted_width = max(adjusted_width, 12)
                elif column_letter in ['K', 'L']:  # Legend columns
                    adjusted_width = max(adjusted_width, 20)
                
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add a summary sheet with just week and quantity
            ws_summary = wb.create_sheet("Přehled")
            
            # Save the file
            filename = f"dodavky_cummins_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=filename
            )
            
            if filepath:
                wb.save(filepath)
                messagebox.showinfo("Hotovo", f"Data byla úspěšně exportována do souboru:\n{filepath}")
                
        except Exception as e:
            messagebox.showerror("Chyba", f"Při exportu došlo k chybě: {str(e)}")
            
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    # When run directly, use the main parser to handle file selection
    from edi_parser_main import EDIUnifiedParser
    EDIUnifiedParser()